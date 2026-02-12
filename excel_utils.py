"""
Excel data management utilities for server/application data
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os
from config import EXCEL_FILE_PATH, EXCEL_SHEET_NAME, COLUMNS


class ExcelDataManager:
    """Handles all Excel operations for server/application data"""
    
    def __init__(self, file_path=EXCEL_FILE_PATH):
        self.file_path = file_path
        self.sheet_name = EXCEL_SHEET_NAME
        self._ensure_file_exists()
    
    def _ensure_file_exists(self):
        """Create Excel file with headers if it doesn't exist"""
        if not os.path.exists(self.file_path):
            os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            
            # Add headers with styling
            headers = list(COLUMNS.values())
            for idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=idx, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, size=12, color="FFFFFF")
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 30
            
            wb.save(self.file_path)
    
    def load_data(self):
        """Load all data from Excel file"""
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb[self.sheet_name]
            
            data = []
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):  # Skip empty rows
                    record = dict(zip(headers, row))
                    data.append(record)
            
            return data
        except Exception as e:
            print(f"Error loading data: {e}")
            return []
    
    def search_by_server(self, server_name):
        """Search for all applications on a specific server"""
        data = self.load_data()
        server_name_lower = server_name.lower()
        
        results = [
            record for record in data
            if record.get(COLUMNS['SERVER_NAME']) and 
            server_name_lower in record[COLUMNS['SERVER_NAME']].lower()
        ]
        
        return results
    
    def search_by_application(self, app_name):
        """Search for all servers running a specific application"""
        data = self.load_data()
        app_name_lower = app_name.lower()
        
        results = [
            record for record in data
            if record.get(COLUMNS['APPLICATION']) and 
            app_name_lower in record[COLUMNS['APPLICATION']].lower()
        ]
        
        return results
    
    def search_all(self, query):
        """Search across all fields"""
        data = self.load_data()
        query_lower = query.lower()
        
        results = []
        for record in data:
            # Search in all text fields
            if any(
                query_lower in str(value).lower()
                for value in record.values()
                if value is not None
            ):
                results.append(record)
        
        return results
    
    def add_record(self, server_name, application, environment="", port="", status="Active", notes=""):
        """Add a new server/application record"""
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb[self.sheet_name]
            
            new_row = [server_name, application, environment, port, status, notes]
            ws.append(new_row)
            
            wb.save(self.file_path)
            return True
        except Exception as e:
            print(f"Error adding record: {e}")
            return False
    
    def get_statistics(self):
        """Get statistics about the data"""
        data = self.load_data()
        
        total_records = len(data)
        unique_servers = len(set(
            record[COLUMNS['SERVER_NAME']] 
            for record in data 
            if record.get(COLUMNS['SERVER_NAME'])
        ))
        unique_apps = len(set(
            record[COLUMNS['APPLICATION']] 
            for record in data 
            if record.get(COLUMNS['APPLICATION'])
        ))
        
        return {
            'total_records': total_records,
            'unique_servers': unique_servers,
            'unique_applications': unique_apps
        }


# Singleton instance
excel_manager = ExcelDataManager()
