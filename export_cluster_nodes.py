"""
Export Cluster Nodes to Excel
Fetches all clusters from Rancher, retrieves nodes for each cluster,
and writes the output to an Excel file.
"""
import sys
import os
from datetime import datetime

# Ensure project root is on the path so we can import config / rancher_utils
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl is required. Install it with:  pip install openpyxl")
    sys.exit(1)

from rancher_utils import rancher_client


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe(value, fallback="N/A"):
    """Return value if truthy, else fallback."""
    if value is None or value == "":
        return fallback
    return value


# ── Main ──────────────────────────────────────────────────────────────────────

def export_nodes_to_excel(output_path=None):
    """Query Rancher for all clusters & nodes and write an Excel workbook."""

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "data",
            f"cluster_nodes_{timestamp}.xlsx",
        )

    # Make sure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print("Connecting to Rancher API …")
    clusters = rancher_client.get_all_clusters()
    print(f"Found {len(clusters)} cluster(s).\n")

    # ── Create workbook ───────────────────────────────────────────────────
    wb = Workbook()

    # ── Styles ────────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    active_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    active_font = Font(color="006100")
    down_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    down_font = Font(color="9C0006")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Sheet 1: Summary (all clusters) ──────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Cluster Summary"

    summary_headers = [
        "Cluster Name",
        "Cluster ID",
        "State",
        "Provider",
        "K8s Version",
        "Total Nodes",
        "Down Nodes",
        "CPU Capacity",
        "CPU Requested",
        "Memory Capacity",
        "Memory Requested",
    ]
    ws_summary.append(summary_headers)

    for col_idx, _ in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Sheet 2: All Nodes ────────────────────────────────────────────────
    ws_nodes = wb.create_sheet("All Nodes")

    node_headers = [
        "Cluster Name",
        "Node Name",
        "State",
        "Roles",
        "OS",
        "Kernel",
        "CPU Count",
        "CPU Capacity",
        "CPU Requested",
        "Memory Capacity",
        "Memory Requested",
        "Allocatable CPU",
        "Allocatable Memory",
    ]
    ws_nodes.append(node_headers)

    for col_idx, _ in enumerate(node_headers, 1):
        cell = ws_nodes.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Populate data ─────────────────────────────────────────────────────
    total_nodes = 0
    for cluster in clusters:
        cid = cluster["id"]
        cname = cluster["name"]
        print(f"  Fetching nodes for cluster: {cname} ({cid}) …")

        try:
            nodes = rancher_client.get_cluster_nodes(cid)
        except Exception as exc:
            print(f"    ⚠  Could not fetch nodes: {exc}")
            nodes = []

        down_count = sum(1 for n in nodes if n.get("is_down"))

        # Summary row
        ws_summary.append([
            cname,
            cid,
            _safe(cluster.get("state")),
            _safe(cluster.get("provider")),
            _safe(cluster.get("k8s_version")),
            len(nodes),
            down_count,
            _safe(cluster.get("cpu_capacity")),
            _safe(cluster.get("cpu_requested")),
            _safe(cluster.get("memory_capacity")),
            _safe(cluster.get("memory_requested")),
        ])

        # Colour-code cluster state
        state_cell = ws_summary.cell(row=ws_summary.max_row, column=3)
        if cluster.get("state", "").lower() == "active":
            state_cell.fill = active_fill
            state_cell.font = active_font
        else:
            state_cell.fill = down_fill
            state_cell.font = down_font

        # Node rows
        for node in nodes:
            ws_nodes.append([
                cname,
                _safe(node.get("name")),
                _safe(node.get("state")),
                ", ".join(node.get("roles", [])) or "N/A",
                _safe(node.get("os_image")),
                _safe(node.get("kernel")),
                _safe(node.get("cpu_count")),
                _safe(node.get("cpu_capacity")),
                _safe(node.get("cpu_requested")),
                _safe(node.get("memory_capacity")),
                _safe(node.get("memory_requested")),
                _safe(node.get("allocatable_cpu")),
                _safe(node.get("allocatable_memory")),
            ])

            # Colour-code node state
            node_state_cell = ws_nodes.cell(row=ws_nodes.max_row, column=3)
            if node.get("state", "").lower() == "active":
                node_state_cell.fill = active_fill
                node_state_cell.font = active_font
            elif node.get("is_down"):
                node_state_cell.fill = down_fill
                node_state_cell.font = down_font

            total_nodes += 1

        print(f"    ✓ {len(nodes)} node(s) ({down_count} down)")

    # ── Auto-fit column widths ────────────────────────────────────────────
    for ws in [ws_summary, ws_nodes]:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ── Add filters ───────────────────────────────────────────────────────
    ws_summary.auto_filter.ref = ws_summary.dimensions
    ws_nodes.auto_filter.ref = ws_nodes.dimensions

    # ── Freeze header row ─────────────────────────────────────────────────
    ws_summary.freeze_panes = "A2"
    ws_nodes.freeze_panes = "A2"

    # ── Save ──────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"Excel report saved → {output_path}")
    print(f"  Clusters: {len(clusters)}  |  Total nodes: {total_nodes}")
    print(f"{'='*60}")
    return output_path


if __name__ == "__main__":
    # Optional: pass a custom output path as first argument
    custom_path = sys.argv[1] if len(sys.argv) > 1 else None
    export_nodes_to_excel(custom_path)
